"""
롯데면세점 SRM 자동 다운로드 - Streamlit 웹앱 버전
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
사용 방법:
  1. 브랜드설정 엑셀 파일을 업로드하세요.
  2. 브랜드 목록을 확인하고 [실행] 버튼을 누르세요.
  3. 완료 후 결과 파일을 ZIP으로 다운로드하세요.
"""

import asyncio
import io
import os
import subprocess
import sys
import tempfile
import traceback
import zipfile
from datetime import datetime

import openpyxl
import streamlit as st
from playwright.async_api import async_playwright


# ── Playwright 브라우저 자동 설치 (서버 환경 대응) ───────────────────────────
@st.cache_resource
def install_playwright_browser():
    result = subprocess.run(
        [sys.executable, "-m", "playwright", "install", "chromium"],
        capture_output=True, text=True
    )
    return result.returncode == 0

# ── 고정 상수 ─────────────────────────────────────────────────────────────────
URL = "https://srm.lottedfs.co.kr/ui/ldfs_ui/index.html"

STORE_CODES = {
    "명동본점": "901",
    "월드타워": "902",
    "부산":     "908",
    "제주":     "90G",
}

STORE_COL_ORDER = ["명동본점", "월드타워", "부산", "제주"]

MD_BTN_SELECTOR = '[id="mainFrame.vFrameSet1.frameTop.form.div_topMenu.form.btn_menu0"]'


# ── 엑셀 설정 읽기 ────────────────────────────────────────────────────────────

def load_brands_from_excel(file_bytes: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet_name = None
    for candidate in ["브랜드설정", "brand_settings", "Sheet", "sheet"]:
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    brands = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        row = list(row) + [None] * 11
        a, b, c, d, e, f, g, h, i, j = (row[k] for k in range(10))

        if not a or not b or not c:
            continue

        store_vals = [d, e, f, g]
        stores = [
            name for name, val in zip(STORE_COL_ORDER, store_vals)
            if str(val or "").strip().upper() == "Y"
        ]
        if not stores:
            continue

        def to_date_str(v):
            if v is None:
                return None
            try:
                return str(int(float(str(v)))).strip()
            except Exception:
                return str(v).strip()

        brands.append({
            "name":      str(a).strip(),
            "user_id":   str(b).strip(),
            "password":  str(c).strip(),
            "stores":    stores,
            "agrg_dvs":  str(h).strip() if h else "Z",
            "date_from": to_date_str(i),
            "date_to":   to_date_str(j),
        })

    wb.close()
    return brands


# ── Playwright 자동화 ─────────────────────────────────────────────────────────

async def wait_js(page, ms):
    await page.wait_for_timeout(ms)


async def safe_wait_for(page, js_check, label, timeout=30000):
    elapsed = 0
    while elapsed < timeout:
        try:
            ok = await page.evaluate(js_check)
            if ok:
                return True
        except Exception:
            pass
        await wait_js(page, 1000)
        elapsed += 1000
    return False


async def do_login(page, user_id, password, log):
    log(f"🔐 로그인 중: {user_id}")
    await page.goto(URL, wait_until="domcontentloaded", timeout=60000)
    await wait_js(page, 4000)
    await page.wait_for_selector("input", timeout=20000)
    await wait_js(page, 1000)

    inputs = await page.query_selector_all("input")
    id_input = pwd_input = None
    for inp in inputs:
        t = (await inp.get_attribute("type")) or "text"
        if t == "password":
            pwd_input = inp
        elif t in ("text", "") and id_input is None:
            id_input = inp

    if not id_input or not pwd_input:
        raise Exception("ID/PW 입력창을 찾을 수 없습니다")

    await id_input.click(click_count=3)
    await id_input.type(user_id, delay=80)
    await pwd_input.click(click_count=3)
    await pwd_input.type(password, delay=80)
    await wait_js(page, 500)
    await pwd_input.press("Enter")
    await wait_js(page, 2000)

    ok = await safe_wait_for(
        page,
        "(() => !!document.querySelector('[id*=\"btn_menu0\"][id*=\"topMenu\"]'))()",
        "로그인 완료",
        timeout=30000
    )
    if not ok:
        raise Exception("로그인 실패 (타임아웃)")
    log("✅ 로그인 성공")
    await wait_js(page, 2000)


async def open_sales_stock_page(page, log):
    log("🔍 매출재고조회 화면 이동 중...")
    await page.click(MD_BTN_SELECTOR)
    await wait_js(page, 2000)

    ok = await safe_wait_for(
        page,
        """
        (() => {
            try {
                var fl = nexacro.getApplication().mainFrame.vFrameSet1.hFrameSet.frameLeft;
                return fl.form.ds_menu.getRowCount() > 0;
            } catch(e) { return false; }
        })()
        """,
        "메뉴 로드", timeout=15000
    )
    if not ok:
        raise Exception("메뉴 데이터 로드 실패")

    nav = await page.evaluate("""
        (() => {
            try {
                var fl  = nexacro.getApplication().mainFrame.vFrameSet1.hFrameSet.frameLeft;
                var ds  = fl.form.ds_menu;
                var grd = fl.form.div_menu.form.grd_menu;
                var cnt = ds.getRowCount();
                var pRow = -1, cRow = -1;
                for (var i = 0; i < cnt; i++) {
                    var nm = ds.getColumn(i, 11) || '';
                    if (nm === '매출재고현황조회') pRow = i;
                    if (nm === '매출재고조회')     cRow = i;
                }
                if (pRow < 0 || cRow < 0)
                    return {ok: false, msg: '메뉴 행 못 찾음', pRow: pRow, cRow: cRow};
                fl.form['div_Menu_grd_Menu_oncellclick'](grd,
                    {row: pRow, col: 0, rowidx: pRow, colidx: 0, button: 1});
                return {ok: true, pRow: pRow, cRow: cRow};
            } catch(e) { return {ok: false, msg: e.toString()}; }
        })()
    """)

    if not nav.get('ok'):
        raise Exception(f"메뉴 탐색 실패: {nav}")

    await wait_js(page, 2000)
    cr = nav['cRow']
    await page.evaluate(f"""
        (() => {{
            var fl  = nexacro.getApplication().mainFrame.vFrameSet1.hFrameSet.frameLeft;
            var grd = fl.form.div_menu.form.grd_menu;
            fl.form['div_Menu_grd_Menu_oncellclick'](grd,
                {{row: {cr}, col: 0, rowidx: {cr}, colidx: 0, button: 1}});
        }})()
    """)

    ok = await safe_wait_for(
        page,
        """
        (() => {
            try {
                var fw  = nexacro.getApplication().mainFrame.vFrameSet1.hFrameSet.vFrameSet2.frameWork;
                var key = Object.keys(fw).find(k => k.startsWith('MO0302'));
                if (!key) return false;
                var sf  = fw[key].form.div_work.form.div_search.form;
                return !!(sf && sf.div_termDate && sf.div_termDate.form && sf.div_termDate.form.cal_start);
            } catch(e) { return false; }
        })()
        """,
        "매출재고조회 화면 로드", timeout=30000
    )
    if not ok:
        raise Exception("매출재고조회 화면 로드 실패")

    log("✅ 매출재고조회 화면 진입")
    await wait_js(page, 1500)


async def download_for_store(page, brand_name, store_name, store_code,
                              agrg_dvs, date_from, date_to, save_dir, log):
    log(f"  📥 [{store_name}] 다운로드 중...")

    set_result = await page.evaluate(f"""
        (() => {{
            try {{
                var fw  = nexacro.getApplication().mainFrame.vFrameSet1.hFrameSet.vFrameSet2.frameWork;
                var key = Object.keys(fw).find(k => k.startsWith('MO0302'));
                var sf  = fw[key].form.div_work.form.div_search.form;

                sf.cbo_strCd.set_value("{store_code}");
                var appliedCode = sf.cbo_strCd.value;
                if (appliedCode !== "{store_code}") {{
                    return {{ok: false, msg: '점 코드 미적용 (해당 점 없음)', tried: "{store_code}", got: appliedCode}};
                }}
                sf.cbo_agrgDvs.set_value("{agrg_dvs}");
                sf.div_termDate.form.cal_start.set_value("{date_from}");
                sf.div_termDate.form.cal_end.set_value("{date_to}");
                return {{
                    ok: true,
                    strCd: sf.cbo_strCd.value,
                    strNm: sf.cbo_strCd.text,
                    dateFrom: sf.div_termDate.form.cal_start.value,
                    dateTo: sf.div_termDate.form.cal_end.value
                }};
            }} catch(e) {{
                return {{ok: false, msg: e.toString()}};
            }}
        }})()
    """)

    if not set_result.get('ok'):
        log(f"  ⚠️ [{store_name}] 조건 설정 실패 (스킵): {set_result.get('msg')}")
        return None

    await wait_js(page, 1000)
    await page.evaluate("""
        (() => {
            var fw  = nexacro.getApplication().mainFrame.vFrameSet1.hFrameSet.vFrameSet2.frameWork;
            var key = Object.keys(fw).find(k => k.startsWith('MO0302'));
            fw[key].form.div_commBtn.form.btn_comSearch.click();
        })()
    """)
    await wait_js(page, 6000)

    store_short = {"명동본점": "명동", "월드타워": "월드타워", "부산": "부산", "제주": "제주"}.get(store_name, store_name)
    yyyymm = date_from[:6]
    safe_brand = brand_name.replace("/", "_")
    fname = f"수불부_{safe_brand}_롯데_{store_short}_{yyyymm}.xlsx"

    try:
        async with page.expect_download(timeout=20000) as dl_info:
            await page.evaluate("""
                (() => {
                    var fw  = nexacro.getApplication().mainFrame.vFrameSet1.hFrameSet.vFrameSet2.frameWork;
                    var key = Object.keys(fw).find(k => k.startsWith('MO0302'));
                    fw[key].form.div_commBtn.form.btn_comExcel.click();
                })()
            """)
        download = await dl_info.value
        save_path = os.path.join(save_dir, fname)
        await download.save_as(save_path)
        log(f"  ✅ [{store_name}] 완료 → {fname}")
        return save_path
    except Exception as e:
        log(f"  ❌ [{store_name}] 다운로드 실패: {e}")
        return None


async def run_automation(brands, save_dir, log, progress_callback):
    results = []
    total_tasks = sum(len(b["stores"]) for b in brands)
    done = 0

    for brand in brands:
        brand_name = brand["name"]
        user_id    = brand["user_id"]
        password   = brand["password"]
        stores     = brand["stores"]
        agrg_dvs   = brand.get("agrg_dvs", "Z")
        date_from  = brand["date_from"]
        date_to    = brand["date_to"]

        if not date_from or not date_to:
            log(f"⚠️ [{brand_name}] 기간 미설정 → 스킵")
            for s in stores:
                results.append({"brand": brand_name, "store": s, "status": "SKIP", "file": None})
                done += 1
                progress_callback(done / total_tasks)
            continue

        log(f"\n{'─'*40}")
        log(f"🏷️ 브랜드: {brand_name}  (ID: {user_id})")
        log(f"   점 목록: {', '.join(stores)}  |  기간: {date_from} ~ {date_to}")

        playwright = await async_playwright().start()
        browser = await playwright.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage",
                  "--disable-blink-features=AutomationControlled"]
        )
        context = await browser.new_context(accept_downloads=True)
        page    = await context.new_page()

        try:
            await do_login(page, user_id, password, log)
            await open_sales_stock_page(page, log)

            for store_name in stores:
                store_code = STORE_CODES.get(store_name)
                if not store_code:
                    log(f"  ⚠️ 알 수 없는 점명: {store_name}")
                    results.append({"brand": brand_name, "store": store_name,
                                    "status": "SKIP", "file": None})
                else:
                    file_path = await download_for_store(
                        page, brand_name, store_name, store_code,
                        agrg_dvs, date_from, date_to, save_dir, log
                    )
                    results.append({
                        "brand":  brand_name,
                        "store":  store_name,
                        "status": "완료" if file_path else "실패",
                        "file":   file_path
                    })
                done += 1
                progress_callback(done / total_tasks)
                await wait_js(page, 1500)

        except Exception:
            err = traceback.format_exc()
            log(f"\n❌ [{brand_name}] 오류 발생:\n{err}")
            for s in stores:
                results.append({"brand": brand_name, "store": s, "status": "오류", "file": None})
                done += 1
                progress_callback(done / total_tasks)
        finally:
            try:
                await browser.close()
            except Exception:
                pass
            await playwright.stop()

    return results


# ── Streamlit UI ──────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="롯데면세점 SRM 자동 다운로드",
        page_icon="🏬",
        layout="centered"
    )

    # 브라우저 설치 (최초 1회만 실행됨)
    with st.spinner("브라우저 준비 중... (최초 실행 시 1~2분 소요)"):
        install_playwright_browser()

    st.title("🏬 롯데면세점 SRM 자동 다운로드")
    st.caption("브랜드설정 엑셀을 업로드하면 SRM에서 매출재고 데이터를 자동으로 내려받습니다.")
    st.divider()

    # ── 파일 업로드 ──────────────────────────────────────────────────────────
    uploaded = st.file_uploader(
        "📂 브랜드설정 엑셀 파일 업로드",
        type=["xlsx"],
        help="'브랜드설정' 시트가 포함된 엑셀 파일을 업로드하세요."
    )

    if not uploaded:
        st.info("엑셀 파일을 업로드하면 브랜드 목록이 표시됩니다.")
        return

    # ── 브랜드 목록 표시 ─────────────────────────────────────────────────────
    try:
        file_bytes = uploaded.read()
        brands = load_brands_from_excel(file_bytes)
    except Exception as e:
        st.error(f"❌ 엑셀 파일 읽기 실패: {e}")
        return

    if not brands:
        st.warning("⚠️ 실행할 브랜드가 없습니다. 엑셀 파일의 내용을 확인하세요.")
        return

    st.success(f"✅ {len(brands)}개 브랜드 로드 완료")

    with st.expander("📋 브랜드 목록 확인", expanded=True):
        for b in brands:
            period = f"{b['date_from']} ~ {b['date_to']}" if b['date_from'] else "⚠️ 기간 미설정"
            st.markdown(
                f"**{b['name']}** &nbsp;·&nbsp; "
                f"ID: `{b['user_id']}` &nbsp;·&nbsp; "
                f"점: {', '.join(b['stores'])} &nbsp;·&nbsp; "
                f"기간: {period}"
            )

    st.divider()

    # ── 실행 버튼 ────────────────────────────────────────────────────────────
    if st.button("▶️ 자동 다운로드 실행", type="primary", use_container_width=True):
        st.session_state["running"] = True

    if not st.session_state.get("running"):
        return

    # ── 실행 영역 ────────────────────────────────────────────────────────────
    log_area     = st.empty()
    progress_bar = st.progress(0, text="준비 중...")
    log_lines    = []

    def log(msg):
        log_lines.append(msg)
        log_area.code("\n".join(log_lines), language=None)

    def progress_callback(ratio):
        pct = int(ratio * 100)
        progress_bar.progress(ratio, text=f"진행 중... {pct}%")

    with tempfile.TemporaryDirectory() as tmpdir:
        log("🚀 실행 시작...")
        try:
            results = asyncio.run(
                run_automation(brands, tmpdir, log, progress_callback)
            )
        except Exception:
            st.error("❌ 실행 중 오류가 발생했습니다.")
            st.code(traceback.format_exc())
            st.session_state["running"] = False
            return

        progress_bar.progress(1.0, text="완료!")

        # ── 결과 요약 ─────────────────────────────────────────────────────
        st.divider()
        st.subheader("📊 결과 요약")

        ok_count   = sum(1 for r in results if r["status"] == "완료")
        fail_count = len(results) - ok_count

        col1, col2 = st.columns(2)
        col1.metric("✅ 성공", f"{ok_count}건")
        col2.metric("❌ 실패/스킵", f"{fail_count}건")

        for r in results:
            icon = "✅" if r["status"] == "완료" else "❌"
            st.markdown(f"{icon} **{r['brand']}** / {r['store']} — {r['status']}")

        # ── ZIP 다운로드 ──────────────────────────────────────────────────
        downloaded_files = [r["file"] for r in results if r["file"]]

        if downloaded_files:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for fpath in downloaded_files:
                    zf.write(fpath, os.path.basename(fpath))
            zip_buffer.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            zip_name  = f"롯데SRM_수불부_{timestamp}.zip"

            st.divider()
            st.download_button(
                label="📦 결과 파일 ZIP 다운로드",
                data=zip_buffer,
                file_name=zip_name,
                mime="application/zip",
                type="primary",
                use_container_width=True
            )
        else:
            st.warning("다운로드된 파일이 없습니다.")

    st.session_state["running"] = False


if __name__ == "__main__":
    main()
